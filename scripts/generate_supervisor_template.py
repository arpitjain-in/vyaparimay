"""
Generate supervisor data-collection template: supervisor_template.xlsx
Run: /usr/local/bin/python3 scripts/generate_supervisor_template.py
Output: scripts/supervisor_template.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ──────────────────────────────────────────────────────────────────
HEADER_BG   = "1F4E79"   # dark blue
HEADER_FG   = "FFFFFF"
FIXED_BG    = "D9D9D9"   # grey  – pre-filled rows (supervisor just enters qty)
INPUT_BG    = "FFFFFF"
REQUIRED_BG = "FFF2CC"   # light yellow – required fields
NOTE_FG     = "7F7F7F"
BORDER_CLR  = "BFBFBF"

thin = Side(style="thin", color=BORDER_CLR)
FULL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_font():
    return Font(name="Calibri", bold=True, color=HEADER_FG, size=11)


def header_fill():
    return PatternFill("solid", fgColor=HEADER_BG)


def fixed_fill():
    return PatternFill("solid", fgColor=FIXED_BG)


def required_fill():
    return PatternFill("solid", fgColor=REQUIRED_BG)


def note_font():
    return Font(name="Calibri", italic=True, color=NOTE_FG, size=9)


def style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = FULL_BORDER


def style_data_row(ws, row, n_cols, fill=None, number_cols=None):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        cell.border = FULL_BORDER
        cell.alignment = Alignment(vertical="center")
        if number_cols and col in number_cols:
            cell.number_format = "#,##0.00"


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = note_font()


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 1 – Instructions
# ─────────────────────────────────────────────────────────────────────────────
def sheet_instructions(wb):
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 90
    ws.row_dimensions[1].height = 36

    title = ws.cell(row=1, column=1, value="📋  Vyaparimay – Supervisor Data Template")
    title.font = Font(name="Calibri", bold=True, size=16, color=HEADER_BG)
    title.alignment = Alignment(vertical="center")

    instructions = [
        "",
        "HOW TO FILL THIS FILE",
        "─────────────────────────────────────────────────────────────────",
        "1.  Fill each sheet using your physical register / stock book.",
        "2.  Yellow cells (🟡) are REQUIRED. Do not leave them blank.",
        "3.  Grey cells are pre-filled — only enter the number in the Value column.",
        "4.  Do NOT change column headers or sheet names.",
        "5.  Use the drop-down lists provided — do not type custom values in those columns.",
        "6.  Dates must be in DD/MM/YYYY format.",
        "7.  All money amounts in ₹ (Indian Rupees). No commas needed.",
        "8.  Save the file and share it back via WhatsApp or email.",
        "",
        "SHEETS IN THIS FILE",
        "─────────────────────────────────────────────────────────────────",
        "  Sheet 2 – Customers        → All customer details from ledger",
        "  Sheet 3 – Invoices         → Historical invoices (one row per line item)",
        "  Sheet 4 – Production Log   → Date-wise milling/production log (kg)",
        "  Sheet 5 – Packaging Stock  → Today's bag/pouch count (pieces)",
        "  Sheet 6 – Ready Stock      → Date-wise packed unit log",
        "  Sheet 7 – Selling Prices   → Current selling rate per unit (₹)",
        "",
        "For any doubt, call: ________________________________",
    ]

    for i, line in enumerate(instructions, 2):
        c = ws.cell(row=i, column=1, value=line)
        if line.startswith("HOW") or line.startswith("SHEETS"):
            c.font = Font(name="Calibri", bold=True, size=11, color=HEADER_BG)
        elif line.startswith("  Sheet"):
            c.font = Font(name="Calibri", size=11)
        elif line.startswith("─"):
            c.font = Font(name="Calibri", color=BORDER_CLR)
        else:
            c.font = Font(name="Calibri", size=11)


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 2 – Customers
# ─────────────────────────────────────────────────────────────────────────────
def sheet_customers(wb):
    ws = wb.create_sheet("Customers")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 40

    headers = [
        "Customer Name *",
        "Firm / Shop Name",
        "Mobile *",
        "Alternate Mobile",
        "Address Line 1 *",
        "Address Line 2",
        "City *",
        "State *",
        "Pin Code",
        "GSTIN",
        "FSSAI No.",
        "Customer Type *",
        "Credit Limit ₹",
        "Payment Terms *",
        "Opening Balance ₹\n(outstanding due)",
        "Notes",
    ]
    n = len(headers)

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(row=1, column=1, value="CUSTOMERS  –  Fill from your ledger register")
    t.font = Font(name="Calibri", bold=True, size=13, color=HEADER_BG)

    # Row 2: legend
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    l = ws.cell(row=2, column=1, value="🟡 Yellow = required   |   * = required field   |   Leave blank if not applicable")
    l.font = note_font()

    # Row 3: headers
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, n)

    # Drop-down validations
    ctype_dv = DataValidation(
        type="list",
        formula1='"Retailer,Wholesaler,Distributor,Direct Consumer"',
        allow_blank=False,
        showErrorMessage=True,
        error="Choose from the list",
        errorTitle="Invalid value",
    )
    pterms_dv = DataValidation(
        type="list",
        formula1='"Cash,7 Days,15 Days,30 Days"',
        allow_blank=False,
    )
    ws.add_data_validation(ctype_dv)
    ws.add_data_validation(pterms_dv)
    ctype_dv.sqref  = f"L4:L200"
    pterms_dv.sqref = f"N4:N200"

    # 30 blank input rows
    req_cols = {1, 3, 5, 7, 8, 12, 14}
    num_cols = {13, 15}
    for row in range(4, 34):
        fill = required_fill() if 1 in req_cols else PatternFill("solid", fgColor=INPUT_BG)
        for col in range(1, n + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = required_fill() if col in req_cols else PatternFill("solid", fgColor=INPUT_BG)
            cell.border = FULL_BORDER
            cell.alignment = Alignment(vertical="center")
            if col in num_cols:
                cell.number_format = "#,##0.00"

    widths = [22, 22, 14, 14, 28, 20, 14, 18, 10, 20, 18, 20, 14, 16, 18, 22]
    set_col_widths(ws, widths)


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 3 – Invoices (one row per line item)
# ─────────────────────────────────────────────────────────────────────────────
def sheet_invoices(wb):
    ws = wb.create_sheet("Invoices")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    n = 13

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(row=1, column=1,
                value="INVOICES  –  One row per line item. Repeat Invoice No + Date for each product on the same bill.")
    t.font = Font(name="Calibri", bold=True, size=13, color=HEADER_BG)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    l = ws.cell(row=2, column=1,
                value="Example: Bill with 2 products = 2 rows with the same Invoice No.  "
                      "Date: DD/MM/YYYY   |   🟡 = required")
    l.font = note_font()

    headers = [
        "Invoice No *\n(e.g. INV-001)",
        "Invoice Date *\n(DD/MM/YYYY)",
        "Customer Name *\n(must match Customers sheet)",
        "Payment Mode *\n(Cash / Credit)",
        "GST Applied?\n(Yes / No)",
        "Product *",
        "Pack Size *",
        "Qty *\n(units)",
        "Rate ₹ *\n(per unit, excl. GST)",
        "Taxable Value ₹\n(auto = Qty × Rate)",
        "CGST ₹",
        "SGST ₹",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, n)
    ws.row_dimensions[3].height = 40

    pmode_dv = DataValidation(
        type="list",
        formula1='"Cash,Credit"',
        allow_blank=False,
        showErrorMessage=True,
        error="Must be Cash or Credit",
        errorTitle="Invalid Payment Mode",
    )
    gst_dv = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
    )
    product_dv = DataValidation(
        type="list",
        formula1='"Shikharji Atta,Shikharji Besan,Shikharji Dalia,Shikharji Bran"',
        allow_blank=False,
    )
    size_dv = DataValidation(
        type="list",
        formula1='"26 kg Bag,5 kg Pouch,10 kg Pouch,5 kg Handle Bag,10 kg Handle Bag,40 kg Bag,500 gm Packet,50 kg Bag,25 kg Bag"',
        allow_blank=False,
    )
    ws.add_data_validation(pmode_dv)
    ws.add_data_validation(gst_dv)
    ws.add_data_validation(product_dv)
    ws.add_data_validation(size_dv)
    pmode_dv.sqref   = "D4:D500"
    gst_dv.sqref     = "E4:E500"
    product_dv.sqref = "F4:F500"
    size_dv.sqref    = "G4:G500"

    req_cols = {1, 2, 3, 4, 6, 7, 8, 9}
    num_cols = {8, 9, 10, 11, 12}
    for row in range(4, 204):  # 200 blank rows
        for col in range(1, n + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill   = required_fill() if col in req_cols else PatternFill("solid", fgColor=INPUT_BG)
            cell.border = FULL_BORDER
            cell.alignment = Alignment(vertical="center")
            if col in num_cols:
                cell.number_format = "#,##0.00"

    set_col_widths(ws, [16, 16, 24, 16, 12, 16, 22, 10, 16, 18, 12, 12, 24])


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 4 – Production Log
# ─────────────────────────────────────────────────────────────────────────────
def sheet_production_log(wb):
    ws = wb.create_sheet("Production Log")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    n = 5

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(row=1, column=1,
                value="PRODUCTION LOG  –  Date-wise milling log  (one row per production batch)")
    t.font = Font(name="Calibri", bold=True, size=13, color=HEADER_BG)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    l = ws.cell(row=2, column=1,
                value="Enter every batch milled from your register — oldest date first.  "
                      "Date format: DD/MM/YYYY   |   🟡 = required")
    l.font = note_font()

    headers = [
        "Date *\n(DD/MM/YYYY)",
        "Time\n(HH:MM, 24hr)",
        "Product *",
        "Quantity Produced *\n(kg)",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, n)
    ws.row_dimensions[3].height = 40

    product_dv = DataValidation(
        type="list",
        formula1='"Shikharji Atta,Shikharji Besan,Shikharji Dalia,Shikharji Bran"',
        allow_blank=False,
        showErrorMessage=True,
        error="Choose from the list",
        errorTitle="Invalid Product",
    )
    ws.add_data_validation(product_dv)
    product_dv.sqref = "C4:C300"

    req_cols = {1, 3, 4}
    for row in range(4, 204):  # 200 blank rows
        for col in range(1, n + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill   = required_fill() if col in req_cols else PatternFill("solid", fgColor=INPUT_BG)
            cell.border = FULL_BORDER
            cell.alignment = Alignment(vertical="center")
            if col == 4:
                cell.number_format = "#,##0.00"

    set_col_widths(ws, [16, 14, 18, 22, 36])


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 5 – Packaging Stock
# ─────────────────────────────────────────────────────────────────────────────
def sheet_packaging_stock(wb):
    ws = wb.create_sheet("Packaging Stock")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="PACKAGING STOCK  –  Count the bags/pouches physically present today")
    t.font = Font(name="Calibri", bold=True, size=13, color=HEADER_BG)

    headers = ["Packaging Material", "Unit", "Quantity (pieces) *"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 3)

    materials = [
        "26 kg Bags – Shikharji Atta",
        "5 kg Pouches – Shikharji Atta",
        "10 kg Pouches – Shikharji Atta",
        "5 kg Handle Bags – Shikharji Atta",
        "10 kg Handle Bags – Shikharji Atta",
        "40 kg Bags – Shikharji Besan",
        "500 gm Packets – Shikharji Besan",
        "500 gm Packets – Shikharji Dalia",
        "50 kg Bags – Shikharji Bran",
        "25 kg Bags – Shikharji Bran",
    ]
    for i, name in enumerate(materials, 3):
        ws.cell(row=i, column=1, value=name).fill = fixed_fill()
        ws.cell(row=i, column=2, value="pcs").fill = fixed_fill()
        qty = ws.cell(row=i, column=3)
        qty.fill = required_fill()
        qty.number_format = "#,##0"
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = FULL_BORDER
            ws.cell(row=i, column=col).alignment = Alignment(vertical="center")

    set_col_widths(ws, [36, 8, 22])


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 5 – Ready Stock (date-wise log)
# ─────────────────────────────────────────────────────────────────────────────
def sheet_ready_stock(wb):
    ws = wb.create_sheet("Ready Stock")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"   # keep header visible while scrolling

    n = 7

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(row=1, column=1,
                value="READY STOCK  –  Date-wise packing log  (one row per packing event)")
    t.font = Font(name="Calibri", bold=True, size=13, color=HEADER_BG)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    l = ws.cell(row=2, column=1,
                value="Enter every packing done from your register — oldest date first. "
                      "Date format: DD/MM/YYYY   |   🟡 = required")
    l.font = note_font()

    headers = [
        "Date *\n(DD/MM/YYYY)",
        "Product *",
        "Pack Size *",
        "Entry Type *\n(ADD / DEDUCT / ADJUST)",
        "Quantity *\n(units)",
        "Reason / Remark *",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, n)
    ws.row_dimensions[3].height = 40

    # Drop-down: product names
    product_dv = DataValidation(
        type="list",
        formula1='"Shikharji Atta,Shikharji Besan,Shikharji Dalia,Shikharji Bran"',
        allow_blank=False,
        showErrorMessage=True,
        error="Choose from the list",
        errorTitle="Invalid Product",
    )
    # Drop-down: pack sizes
    size_dv = DataValidation(
        type="list",
        formula1='"26 kg Bag,5 kg Pouch,10 kg Pouch,5 kg Handle Bag,10 kg Handle Bag,40 kg Bag,500 gm Packet,50 kg Bag,25 kg Bag"',
        allow_blank=False,
    )
    # Drop-down: entry type
    type_dv = DataValidation(
        type="list",
        formula1='"ADD,DEDUCT,ADJUST"',
        allow_blank=False,
        showErrorMessage=True,
        error="Must be ADD, DEDUCT or ADJUST",
        errorTitle="Invalid Type",
    )
    ws.add_data_validation(product_dv)
    ws.add_data_validation(size_dv)
    ws.add_data_validation(type_dv)
    product_dv.sqref = "B4:B300"
    size_dv.sqref    = "C4:C300"
    type_dv.sqref    = "D4:D300"

    # 100 blank input rows
    req_cols = {1, 2, 3, 4, 5, 6}
    for row in range(4, 104):
        for col in range(1, n + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill   = required_fill() if col in req_cols else PatternFill("solid", fgColor=INPUT_BG)
            cell.border = FULL_BORDER
            cell.alignment = Alignment(vertical="center")
            if col == 5:
                cell.number_format = "#,##0"

    set_col_widths(ws, [16, 16, 22, 22, 12, 28, 24])


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 6 – Selling Prices
# ─────────────────────────────────────────────────────────────────────────────
def sheet_prices(wb):
    ws = wb.create_sheet("Selling Prices")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="SELLING PRICES  –  Current rate charged to customers (₹ per unit, excluding GST)")
    t.font = Font(name="Calibri", bold=True, size=13, color=HEADER_BG)

    headers = ["Product / SKU", "Pack Size", "Rate ₹ per unit *"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 3)

    skus = [
        ("Shikharji Atta",  "26 kg Bag",        780),
        ("Shikharji Atta",  "5 kg Pouch",        165),
        ("Shikharji Atta",  "10 kg Pouch",       320),
        ("Shikharji Atta",  "5 kg Handle Bag",   175),
        ("Shikharji Atta",  "10 kg Handle Bag",  340),
        ("Shikharji Besan",        "40 kg Bag",        2400),
        ("Shikharji Besan",        "500 gm Packet",      35),
        ("Shikharji Dalia",       "500 gm Packet",      28),
        ("Shikharji Bran",         "50 kg Bag",         600),
        ("Shikharji Bran",         "25 kg Bag",         310),
    ]
    for i, (product, size, default_rate) in enumerate(skus, 3):
        ws.cell(row=i, column=1, value=product).fill  = fixed_fill()
        ws.cell(row=i, column=2, value=size).fill     = fixed_fill()
        rate = ws.cell(row=i, column=3, value=default_rate)
        rate.fill = required_fill()
        rate.number_format = "₹#,##0.00"
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = FULL_BORDER
            ws.cell(row=i, column=col).alignment = Alignment(vertical="center")

    ws.merge_cells("A14:C14")
    note = ws.cell(row=14, column=1,
                   value="NOTE: Default rates are pre-filled from the system. Update any that have changed.")
    note.font = note_font()

    set_col_widths(ws, [22, 22, 22])


# ─────────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    sheet_instructions(wb)
    sheet_customers(wb)
    sheet_invoices(wb)
    sheet_production_log(wb)
    sheet_packaging_stock(wb)
    sheet_ready_stock(wb)
    sheet_prices(wb)

    out = "scripts/supervisor_template.xlsx"
    wb.save(out)
    print(f"✅  Template saved → {out}")


if __name__ == "__main__":
    main()
